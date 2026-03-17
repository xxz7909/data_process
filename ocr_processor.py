"""
股票图片OCR批量识别工具
功能：批量识别 差额/代码/市值 图片，提取关键信息写入Excel
依赖：rapidocr-onnxruntime, pillow, pandas, openpyxl
"""
import os
import re
import logging
import numpy as np
from PIL import Image
from rapidocr_onnxruntime import RapidOCR
import pandas as pd


def create_ocr():
    """创建RapidOCR实例（基于PP-OCRv4 ONNX模型）"""
    return RapidOCR()


def load_image(path, max_width=1200):
    """加载图片并限制宽度以加速OCR，返回RGB numpy数组"""
    img = Image.open(path)
    if img.width > max_width:
        ratio = max_width / img.width
        img = img.resize((max_width, int(img.height * ratio)), Image.LANCZOS)
    if img.mode != 'RGB':
        img = img.convert('RGB')
    return np.array(img)


def ocr_image(ocr_engine, path):
    """OCR识别图片，返回文本行列表"""
    img_arr = load_image(path)
    result, _ = ocr_engine(img_arr)
    if result:
        return [item[1] for item in result]
    return []


def ocr_image_with_boxes(ocr_engine, path):
    """OCR识别图片，返回 [(text, box, confidence), ...] 及原始图片数组"""
    img_arr = load_image(path)
    result, _ = ocr_engine(img_arr)
    items = []
    if result:
        for item in result:
            box = item[0]          # [[x1,y1],[x2,y2],[x3,y3],[x4,y4]]
            text = item[1]
            conf = item[2]
            items.append((text, box, conf))
    return items, img_arr


def _normalize_minus(text):
    """将各种减号/破折号统一为半角减号"""
    for ch in '一—–−﹣':
        text = text.replace(ch, '-')
    # 全角减号
    text = text.replace('\uff0d', '-')
    return text


def _detect_text_color(img_arr, box):
    """
    检测文本框内文字的主色调，判断红/绿。
    返回: 'red', 'green', 或 'unknown'
    """
    pts = np.array(box, dtype=np.int32)
    x_min, y_min = pts.min(axis=0)
    x_max, y_max = pts.max(axis=0)
    # 裁剪区域（边界保护）
    h, w = img_arr.shape[:2]
    x_min, y_min = max(0, x_min), max(0, y_min)
    x_max, y_max = min(w, x_max), min(h, y_max)
    crop = img_arr[y_min:y_max, x_min:x_max]
    if crop.size == 0:
        return 'unknown'

    # 转到float处理
    pixels = crop.reshape(-1, 3).astype(np.float32)
    r, g, b = pixels[:, 0], pixels[:, 1], pixels[:, 2]

    # 过滤掉背景色（接近白色或接近黑色的像素）
    brightness = (r + g + b) / 3
    mask = (brightness > 30) & (brightness < 230)
    if mask.sum() < 10:
        return 'unknown'
    r, g, b = r[mask], g[mask], b[mask]

    # 红色像素：R 通道显著高于 G 和 B
    red_mask = (r > 120) & (r > g * 1.4) & (r > b * 1.4)
    # 绿色像素：G 通道显著高于 R 和 B
    green_mask = (g > 120) & (g > r * 1.4) & (g > b * 1.4)

    red_ratio = red_mask.sum() / len(r)
    green_ratio = green_mask.sum() / len(r)

    if green_ratio > 0.05 and green_ratio > red_ratio:
        return 'green'
    elif red_ratio > 0.05 and red_ratio > green_ratio:
        return 'red'
    return 'unknown'


def extract_diff(lines, ocr_items=None, img_arr=None):
    """从差额图片中提取差额(亿)数值，支持正负号检测。

    OCR文本格式示例：
      "亿：-14.12差额亿：90.55 著名游资低"
      "-5.35差额亿：16.37著名游资低吸：0"
    其中第一个数值（可能为负）是差额，"差额亿：XX"后的值是另一个指标。
    """
    text = ''.join(lines)
    text = _normalize_minus(text)

    # 策略：提取文本中**第一个**浮点数（可能带负号），这就是差额值
    match = re.search(r'(-?\d+\.?\d*)', text)
    if not match:
        return ''

    value_str = match.group(1)
    has_text_minus = value_str.startswith('-')

    # 颜色检测仅作为补充：当OCR文本没有负号时，用颜色判断是否为负值
    # 当OCR文本已有负号时，信任OCR结果（不让颜色覆盖）
    if not has_text_minus and ocr_items and img_arr is not None:
        color = 'unknown'
        for item_text, box, _ in ocr_items:
            normalized = _normalize_minus(item_text)
            if value_str in normalized:
                color = _detect_text_color(img_arr, box)
                if color != 'unknown':
                    break
        # 绿色=负值，补加负号
        if color == 'green':
            value_str = '-' + value_str

    return value_str


def _is_valid_stock_code(code_str):
    """检查是否为合法A股代码前缀"""
    valid_prefixes = [
        '000', '001', '002', '003',          # 深市主板/中小板
        '300', '301',                         # 创业板
        '600', '601', '603', '605',           # 沪市主板
        '688',                                # 科创板
        '830', '831', '832', '833', '834',    # 北交所
        '835', '836', '837', '838', '839',
        '920',                                # 北交所
    ]
    return any(code_str.startswith(p) for p in valid_prefixes)


def extract_code_and_name(lines):
    """从代码图片中提取6位代码和汉字名称"""
    full_text = ' '.join(lines)

    # 提取6位数字代码 —— 优先匹配合法股票代码前缀
    code = ''
    # 找所有连续数字段
    digit_segments = re.findall(r'\d+', full_text)
    for seg in digit_segments:
        if len(seg) == 6 and _is_valid_stock_code(seg):
            code = seg
            break
        elif len(seg) > 6:
            # 从长数字段中滑窗找合法6位代码
            for j in range(len(seg) - 5):
                candidate = seg[j:j + 6]
                if _is_valid_stock_code(candidate):
                    code = candidate
                    break
            if code:
                break
    # 回退：匹配任意6位数字
    if not code:
        code_match = re.search(r'(\d{6})', full_text)
        if code_match:
            code = code_match.group(1)
    # 再回退：匹配5位数字（OCR可能截断）
    if not code:
        code_match = re.search(r'(\d{5})', full_text)
        if code_match:
            code = code_match.group(1)

    # 提取股票名称（中文字符，可能带 -U, -UW 等后缀）
    # 排除常见干扰词（含UI界面噪声文字）
    noise_words = {'决策', '传统', '交易', '亮点', '特色', '涨停', '跌停',
                   '盘面', '资金', '主力', '散户', '买入', '卖出', '成交',
                   '板块', '行情', '分时', '日线', '周线', '月线', '指标'}
    # OCR末尾常见噪声字符（仅中文单字噪声）
    trailing_noise = set('高托拉上下')
    name = ''
    for line in lines:
        line_clean = line.strip()
        if line_clean in noise_words:
            continue
        # 匹配以中文开头，可能有 -U/-UW 后缀的股票名
        name_match = re.match(r'([\u4e00-\u9fa5]{2,}(?:-[A-Za-z]{1,3})?)', line_clean)
        if name_match:
            candidate = name_match.group(1)
            if candidate not in noise_words and len(candidate) >= 2:
                # 仅去除末尾中文噪声字符（不影响英文后缀）
                while candidate and candidate[-1] in trailing_noise:
                    candidate = candidate[:-1]
                if len(candidate) >= 2:
                    name = candidate
                    break

    return name, code


def extract_market_cap(lines):
    """从市值图片中提取总市值（纯数值，已去除'亿'和逗号）"""
    full_text = '\n'.join(lines)

    # 找到所有带"亿"的数值
    values = re.findall(r'([\d,.]+)亿', full_text)
    if not values:
        return ''

    # 总市值是最大的那个亿值
    max_val = ''
    max_num = 0
    for v in values:
        num_str = v.replace(',', '')
        try:
            num = float(num_str)
            if num > max_num:
                max_num = num
                max_val = num_str
        except ValueError:
            continue

    return max_val


def detect_groups(img_dir):
    """自动检测图片组数"""
    groups = set()
    if not os.path.isdir(img_dir):
        return 0
    for fname in os.listdir(img_dir):
        match = re.match(r'^(\d+)_', fname)
        if match:
            groups.add(int(match.group(1)))
    return max(groups) if groups else 0


def process_images(img_dir, output_path, progress_callback=None):
    """
    主处理函数：批量识别图片并写入Excel
    
    Args:
        img_dir: 图片目录路径
        output_path: 输出Excel路径
        progress_callback: 进度回调函数 callback(current, total, message)
    
    Returns:
        (success, message) 元组
    """
    group_count = detect_groups(img_dir)
    if group_count == 0:
        return False, "未在指定目录中找到图片文件（格式：N_差额.png.bmp）"

    ocr_engine = create_ocr()
    results = []
    total_steps = group_count * 3  # 每组3张图

    for i in range(1, group_count + 1):
        step_base = (i - 1) * 3

        # --- 差额（使用带坐标的OCR结果+颜色检测） ---
        diff_path = os.path.join(img_dir, f'{i}_差额.png.bmp')
        diff_value = ''
        if os.path.exists(diff_path):
            if progress_callback:
                progress_callback(step_base + 1, total_steps, f'正在识别 第{i}组 差额...')
            try:
                ocr_items, img_arr = ocr_image_with_boxes(ocr_engine, diff_path)
                lines = [item[0] for item in ocr_items]
                diff_value = extract_diff(lines, ocr_items, img_arr)
            except Exception as e:
                diff_value = f'识别失败: {e}'

        # --- 代码 ---
        code_path = os.path.join(img_dir, f'{i}_代码.png.bmp')
        name, code = '', ''
        if os.path.exists(code_path):
            if progress_callback:
                progress_callback(step_base + 2, total_steps, f'正在识别 第{i}组 代码...')
            try:
                lines = ocr_image(ocr_engine, code_path)
                name, code = extract_code_and_name(lines)
            except Exception as e:
                name = f'识别失败: {e}'

        # --- 市值 ---
        value_path = os.path.join(img_dir, f'{i}_市值.png.bmp')
        market_cap = ''
        if os.path.exists(value_path):
            if progress_callback:
                progress_callback(step_base + 3, total_steps, f'正在识别 第{i}组 市值...')
            try:
                lines = ocr_image(ocr_engine, value_path)
                market_cap = extract_market_cap(lines)
            except Exception as e:
                market_cap = f'识别失败: {e}'

        results.append({
            '序号': i,
            '差额(亿)': diff_value,
            '名称': name,
            '代码': code,
            '总市值': market_cap,
        })

    # 写入Excel
    df = pd.DataFrame(results)
    df.to_excel(output_path, index=False, engine='openpyxl')

    # 调整列宽
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    ws = wb.active
    col_widths = {'A': 8, 'B': 14, 'C': 18, 'D': 12, 'E': 16}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    # 表头加粗
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)
    wb.save(output_path)

    return True, f"处理完成！共 {group_count} 组数据已写入 {output_path}"


if __name__ == '__main__':
    # 命令行模式直接运行
    img_dir = 'pics'
    output_path = 'result.xlsx'

    def cli_progress(current, total, msg):
        print(f'[{current}/{total}] {msg}')

    success, message = process_images(img_dir, output_path, cli_progress)
    print(message)
